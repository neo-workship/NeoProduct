import pandas as pd
import json
import random
from openpyxl import load_workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta
from typing import List, Dict, Any, Optional
import hashlib

class FlatEnterpriseArchiveGenerator:
    """
    扁平化企业档案生成器
    将Excel的层级结构转换为扁平化的MongoDB文档集合
    """
    
    def __init__(self):
        self.level_counters = {
            'l1': {},  # 一级分类计数器
            'l2': {},  # 二级分类计数器  
            'l3': {},  # 三级分类计数器
            'field': {}  # 字段计数器
        }
        
    def generate_code(self, name: str, level: str, parent_code: str = "") -> str:
        """
        生成层级编码
        Args:
            name: 分类或字段名称
            level: 层级类型 (l1, l2, l3, field)
            parent_code: 父级编码
        Returns:
            str: 生成的编码
        """
        # 使用MD5哈希确保编码唯一性和一致性
        hash_input = f"{parent_code}_{name}" if parent_code else name
        hash_code = hashlib.md5(hash_input.encode('utf-8')).hexdigest()[:6].upper()
        
        if level == 'l1':
            return f"L1{hash_code}"
        elif level == 'l2':
            return f"L2{hash_code}"
        elif level == 'l3':
            return f"L3{hash_code}"
        elif level == 'field':
            return f"F{hash_code}"
        
        return hash_code

    def read_excel_with_merged_cells(self, file_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        读取包含合并单元格的Excel文件并转换为扁平化文档列表
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称，默认为None（第一个工作表）
        Returns:
            List[Dict]: 扁平化的字段文档列表
        """
        # 使用openpyxl读取Excel文件以处理合并单元格
        workbook = load_workbook(file_path)
        if sheet_name is None:
            worksheet = workbook.active
        else:
            worksheet = workbook[sheet_name]
            
        # 获取所有数据
        data = []
        for row in worksheet.iter_rows(values_only=True):
            data.append(list(row))
            
        # 处理合并单元格 - 填充合并单元格的值
        merged_ranges = worksheet.merged_cells.ranges
        for merged_range in merged_ranges:
            # 获取合并区域的值（左上角单元格的值）
            merge_value = worksheet[merged_range.coord.split(':')[0]].value 
            # 填充合并区域内的所有单元格
            for row_idx in range(merged_range.min_row - 1, merged_range.max_row):
                for col_idx in range(merged_range.min_col - 1, merged_range.max_col):
                    if row_idx < len(data) and col_idx < len(data[row_idx]):
                        data[row_idx][col_idx] = merge_value
                        
        # 转换为扁平化文档列表
        flat_documents = []
        current_time = datetime.now()
        
        # 用于跟踪当前的层级信息
        current_l1 = {"name": "", "code": ""}
        current_l2 = {"name": "", "code": ""}
        current_l3 = {"name": "", "code": ""}
        
        # 用于排序
        l1_order = 0
        l2_order = 0
        l3_order = 0
        
        # 假设表格从第2行开始（第1行是标题）
        for row_idx, row_data in enumerate(data[1:], start=2):
            if len(row_data) < 5:  # 确保至少有5列数据
                continue
                
            # 提取各级分类数据
            level1 = row_data[1] if row_data[1] is not None else ""
            level2 = row_data[2] if row_data[2] is not None else ""
            level3 = row_data[3] if row_data[3] is not None else ""
            field_name = row_data[4] if row_data[4] is not None else ""
            remark = row_data[5] if len(row_data) > 5 and row_data[5] is not None else ""
            data_url = row_data[6] if len(row_data) > 6 and row_data[6] is not None else ""
            
            # 跳过空行
            if not any([level1, level2, level3, field_name]):
                continue
                
            # 更新层级信息和编码
            if level1 and level1 != current_l1["name"]:
                l1_order += 1
                l2_order = 0
                l3_order = 0
                current_l1 = {
                    "name": level1,
                    "code": self.generate_code(level1, "l1")
                }
                
            if level2 and level2 != current_l2["name"]:
                l2_order += 1
                l3_order = 0
                current_l2 = {
                    "name": level2,
                    "code": self.generate_code(level2, "l2", current_l1["code"])
                }
                
            if level3 and level3 != current_l3["name"]:
                l3_order += 1
                current_l3 = {
                    "name": level3,
                    "code": self.generate_code(level3, "l3", current_l2["code"])
                }
            
            # 生成字段编码
            field_code = self.generate_code(field_name, "field", current_l3["code"])
            
            # 构建路径
            path_code = f"{current_l1['code']}.{current_l2['code']}.{current_l3['code']}"
            path_name = f"{current_l1['name']}.{current_l2['name']}.{current_l3['name']}"
            full_path_code = f"{path_code}.{field_code}"
            full_path_name = f"{path_name}.{field_name}"
            
            # 生成扁平化文档
            field_document = {
                # 企业信息（稍后填充）
                "enterprise_code": "",
                "enterprise_name": "",
                
                # 层级信息
                "l1_code": current_l1["code"],
                "l1_name": current_l1["name"],
                "l2_code": current_l2["code"],
                "l2_name": current_l2["name"],
                "l3_code": current_l3["code"],
                "l3_name": current_l3["name"],
                
                # 路径信息
                "path_code": path_code,
                "path_name": path_name,
                "full_path_code": full_path_code,
                "full_path_name": full_path_name,
                
                # 字段信息
                "field_code": field_code,
                "field_name": field_name,
                "field_type": self._infer_field_type(field_name),
                
                # 数据值
                "value": "",  # 初始值为空
                "value_text": "",  # 用于全文检索的文本值
                
                # 元数据
                "remark": remark,
                "data_url": data_url,
                "is_required": self._is_required_field(field_name, remark),
                "data_source": self._extract_data_source(remark),
                "encoding": "UTF-8",
                "format": "text/plain",
                "license": "企业数据使用许可",
                "rights": "企业内部使用，受数据保护法规约束",
                "update_frequency":"更新频率",
                "value_dict":{"xx","xx","xx"},
                
                # 排序和显示
                "l1_order": l1_order,
                "l2_order": l2_order,
                "l3_order": l3_order,
                "field_order": len(flat_documents) + 1,
                
                # 时间戳
                "create_time": current_time,
                "update_time": current_time,
                
                # 数据状态
                "status": "active"
            }
            
            flat_documents.append(field_document)
            
        return flat_documents
    
    def _infer_field_type(self, field_name: str) -> str:
        """
        根据字段名推断字段类型
        """
        field_name_lower = field_name.lower()
        
        if any(keyword in field_name_lower for keyword in ['金额', '资金', '资本', "总额", '收入', '支出', '数量', '面积']):
            return "number"
        elif any(keyword in field_name_lower for keyword in ['时间', '日期', '年份', '月份']):
            return "date"
        elif any(keyword in field_name_lower for keyword in ['是否', '有无']):
            return "boolean"
        elif any(keyword in field_name_lower for keyword in ['邮箱', '邮件']):
            return "email"
        elif any(keyword in field_name_lower for keyword in ['电话', '手机', '传真']):
            return "phone"
        elif any(keyword in field_name_lower for keyword in ['网址', '网站', 'url']):
            return "url"
        else:
            return "string"
    
    def _is_required_field(self, field_name: str, remark: str) -> bool:
        """
        判断字段是否为必填项
        """
        required_keywords = ['必填', '必须', '必需', '必要', '必备']
        field_text = f"{field_name} {remark}".lower()
        return any(keyword in field_text for keyword in required_keywords)
    
    def _extract_data_source(self, remark: str) -> str:
        """
        从备注中提取数据来源
        """
        if not remark:
            return ""
            
        sources = ['工商局', '税务局', '银行', '统计局', '人社局', '公安局', '海关']
        for source in sources:
            if source in remark:
                return source
        return ""

    def generate_enterprise_documents(self, template_documents: List[Dict[str, Any]], 
                                   enterprise_code: str, 
                                   enterprise_name: str,
                                   field_values: Optional[Dict[str, str]] = None) -> List[Dict[str, Any]]:
        """
        为特定企业生成完整的字段文档集合
        Args:
            template_documents: 模板文档列表
            enterprise_code: 企业代码
            enterprise_name: 企业名称
            field_values: 字段值字典 {field_name: value}
        Returns:
            List[Dict]: 企业的完整字段文档集合
        """
        enterprise_documents = []
        current_time = datetime.now()
        
        for template_doc in template_documents:
            # 复制模板文档
            enterprise_doc = template_doc.copy()
            
            # 设置企业信息
            enterprise_doc["enterprise_code"] = enterprise_code
            enterprise_doc["enterprise_name"] = enterprise_name
            
            # 设置字段值
            field_name = template_doc["field_name"]
            if field_values and field_name in field_values:
                value = field_values[field_name]
                enterprise_doc["value"] = value
                enterprise_doc["value_text"] = str(value)  # 用于全文检索
                enterprise_doc["update_time"] = current_time
            
            enterprise_documents.append(enterprise_doc)
            
        return enterprise_documents

    def generate_sample_data(self, template_documents: List[Dict[str, Any]], 
                           num_enterprises: int = 5) -> List[Dict[str, Any]]:
        """
        生成示例企业数据
        """
        all_documents = []
        
        sample_enterprises = [
            {"code": "E001", "name": "示例科技有限公司"},
            {"code": "E002", "name": "创新制造股份有限公司"},
            {"code": "E003", "name": "绿色能源集团"},
            {"code": "E004", "name": "智能物流有限公司"},
            {"code": "E005", "name": "数字金融科技公司"}
        ]
        
        # 示例字段值
        sample_values = {
            "统一社会信用代码": ["91110000123456789X", "91120000234567890Y", "91130000345678901Z"],
            "企业名称": ["示例科技有限公司", "创新制造股份有限公司", "绿色能源集团"],
            "注册资金": ["1000万元", "5000万元", "2000万元", "800万元", "3000万元"],
            "企业类型": ["有限责任公司", "股份有限公司", "国有企业"],
            "行业分类": ["软件和信息技术服务业", "制造业", "新能源", "物流业", "金融业"]
        }
        
        for i, enterprise in enumerate(sample_enterprises[:num_enterprises]):
            # 为每个字段生成示例值
            field_values = {}
            for template_doc in template_documents:
                field_name = template_doc["field_name"]
                if field_name in sample_values:
                    field_values[field_name] = sample_values[field_name][i % len(sample_values[field_name])]
                elif template_doc["field_type"] == "number":
                    field_values[field_name] = str(random.randint(100, 10000))
                elif template_doc["field_type"] == "date":
                    field_values[field_name] = datetime.now().strftime("%Y-%m-%d")
                elif template_doc["field_type"] == "boolean":
                    field_values[field_name] = random.choice(["是", "否"])
                else:
                    field_values[field_name] = f"示例{field_name}值"
            
            # 生成企业文档
            enterprise_docs = self.generate_enterprise_documents(
                template_documents, 
                enterprise["code"], 
                enterprise["name"],
                field_values
            )
            all_documents.extend(enterprise_docs)
            
        return all_documents

def main():
    """
    主函数：演示扁平化文档生成流程
    """
    # 创建生成器实例
    generator = FlatEnterpriseArchiveGenerator()
    
    # Excel文件路径
    file_path = "./一企一档数据项.xlsx"
    
    try:
        print("开始读取Excel文件...")
        # 读取Excel并生成模板文档
        template_documents = generator.read_excel_with_merged_cells(file_path)
        print(f"成功生成 {len(template_documents)} 个字段模板")
        
        # 保存结果
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        # 生成示例企业数据
        # print("生成示例企业数据...")
        # sample_documents = generator.generate_sample_data(template_documents, num_enterprises=3)
        # print(f"成功生成 {len(sample_documents)} 个企业字段文档")
        # sample_file = f"示例企业数据_{current_time}.json"
        # with open(sample_file, 'w', encoding='utf-8') as f:
        #     json.dump(sample_documents, f, ensure_ascii=False, indent=2, default=str)
        # print(f"示例企业数据已保存为: {sample_file}")
        # print(f"\n示例企业数量: {len(sample_documents) // len(template_documents)}")
        
        # 保存模板文档
        template_file = f"字段模板_{current_time}.json"
        with open(template_file, 'w', encoding='utf-8') as f:
            json.dump(template_documents, f, ensure_ascii=False, indent=2, default=str)
        print(f"字段模板已保存为: {template_file}")
   
        # 打印统计信息
        print("\n=== 数据统计 ===")
        print(f"总字段数: {len(template_documents)}")
        
        # 按层级统计
        l1_stats = {}
        for doc in template_documents:
            l1_name = doc["l1_name"]
            if l1_name not in l1_stats:
                l1_stats[l1_name] = 0
            l1_stats[l1_name] += 1
        
        print("\n一级分类统计:")
        for l1_name, count in l1_stats.items():
            print(f"  {l1_name}: {count} 个字段")
            
        # 字段类型统计
        type_stats = {}
        for doc in template_documents:
            field_type = doc["field_type"]
            if field_type not in type_stats:
                type_stats[field_type] = 0
            type_stats[field_type] += 1
            
        print("\n字段类型统计:")
        for field_type, count in type_stats.items():
            print(f"  {field_type}: {count} 个字段")
            
        print(f"每个企业字段数: {len(template_documents)}")
        
    except FileNotFoundError:
        print(f"错误: 找不到文件 {file_path}")
        print("请确保Excel文件存在且路径正确")
    except Exception as e:
        print(f"错误: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()