import pandas as pd
import json
import random
from openpyxl import load_workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta

def read_excel_with_merged_cells(file_path, sheet_name=None, selected_enterprise_code=None, selected_enterprise_name=None):
    """
    读取包含合并单元格的Excel文件并转换为4级JSON结构
    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称，默认为None（第一个工作表）
        selected_enterprise_code: 选定的企业代码，用于同步
        selected_enterprise_name: 选定的企业名称，用于同步
    Returns:
        dict: 4级嵌套的JSON结构
    """
    # 使用openpyxl读取Excel文件以处理合并单元格
    workbook = load_workbook(file_path)
    if sheet_name is None:
        worksheet = workbook.active
    else:
        worksheet = workbook[sheet_name]
    # 获取所有数据
    data = []
    for row in worksheet.iter_rows(values_only=True):
        data.append(list(row))
    # 处理合并单元格 - 填充合并单元格的值
    merged_ranges = worksheet.merged_cells.ranges
    for merged_range in merged_ranges:
        # 获取合并区域的值（左上角单元格的值）
        merge_value = worksheet[merged_range.coord.split(':')[0]].value 
        # 填充合并区域内的所有单元格
        for row_idx in range(merged_range.min_row - 1, merged_range.max_row):
            for col_idx in range(merged_range.min_col - 1, merged_range.max_col):
                if row_idx < len(data) and col_idx < len(data[row_idx]):
                    data[row_idx][col_idx] = merge_value
    # 构建4级JSON结构
    result = {}
    # 假设表格从第2行开始（第1行是标题）
    for row_idx, row_data in enumerate(data[1:], start=2):  # 跳过标题行
        if len(row_data) < 6:  # 确保至少有6列数据
            continue     
        # 提取各级分类数据
        level1 = row_data[1] if row_data[1] is not None else ""  # 一级分类
        level2 = row_data[2] if row_data[2] is not None else ""  # 二级分类
        level3 = row_data[3] if row_data[3] is not None else ""  # 三级分类
        level4 = row_data[4] if row_data[4] is not None else ""  # 展示字段
        remark = row_data[5] if len(row_data) > 6 and row_data[5] is not None else ""  # 备注
        data_url = row_data[6] if len(row_data) > 6 and row_data[6] is not None else ""  # 备注

        # 跳过空行
        if not any([level1, level2, level3, level4]):
            continue    
        # 构建层级结构
        if level1 not in result:
            result[level1] = {}   
        if level2 not in result[level1]:
            result[level1][level2] = {}
        if level3 not in result[level1][level2]:
            result[level1][level2][level3] = []
        
        # 生成第四级数据（展示字段），传入选定的企业代码和名称进行同步
        field_info = generate_field_info(level4, remark, data_url, selected_enterprise_code, selected_enterprise_name)
        result[level1][level2][level3].append(field_info)
    return result

def generate_field_info(field_name, remark, data_url, selected_enterprise_code, selected_enterprise_name):
    """
    生成字段信息，包含data和meta部分，并同步测试数据。
    """
    today = datetime.now().strftime("%Y-%m-%d")
    field_info = {
        "field_name": field_name,
        "type":"level4",
        "remark": remark,
        "data": {
            "value": "",  # 设置为空字符串
            "data_url":data_url
        },
    }
    return field_info

def convert_to_enterprise_json(data, enterprise_code=None, enterprise_name=None):
    """
    转换为企业为主文档的JSON结构，并添加各级meta信息
    """
    children = []
    today =datetime.now().strftime("%Y-%m-%d")

    for level1_name, level1_data in data.items():
        level1_children = []
        for level2_name, level2_data in level1_data.items():
            level2_children = []
            for level3_name, level3_data in level2_data.items():
                level3_item = {
                    "name": level3_name,
                    "type": "level3",
                    "fields": level3_data
                }
                level2_children.append(level3_item)

            level2_item = {
                "name": level2_name,
                "type": "level2",
                "children": level2_children
            }
            level1_children.append(level2_item)

        level1_item = {
            "name": level1_name,
            "type": "level1",
            "children": level1_children
        }
        children.append(level1_item)

    return {
        "EnterpriseCode": enterprise_code,
        "EnterpriseName": enterprise_name,
        "type": "level0",
        "CreateTime": today,
        "children": children
    }

if __name__ ==  "__main__":
    file_path = "./一企一档数据项.xlsx"
    json_data=read_excel_with_merged_cells(file_path)
    enterprise_json=convert_to_enterprise_json(json_data)
    #保存到文件（可选）
    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file = f"企业档案_结构化_{current_time}.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(enterprise_json, f, ensure_ascii=False, indent=2)
    print(f"转换完成！文件已保存为: {output_file}")
