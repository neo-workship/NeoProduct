import pandas as pd
import json
import random
from openpyxl import load_workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta

# 加载测试数据 (不再加载，设置为空)
test_data = {}
test_enterprises = []

def get_current_date():
    years_to_subtract = random.randint(5, 8)
    past_date = datetime.now() # - relativedelta(years=years_to_subtract)
    return past_date.strftime("%Y-%m-%d")
    # return datetime.now().strftime("%Y-%m-%d")

def get_test_value(field_name, test_data, selected_enterprise_code=None, selected_enterprise_name=None):
    """
    根据字段名获取测试数据。
    如果字段是“统一社会信用代码”或“企业名称”，则使用传入的selected_enterprise_code/name。
    否则，返回空字符串。
    """
    if field_name == "统一社会信用代码":
        return selected_enterprise_code if selected_enterprise_code else ""
    elif field_name == "企业名称":
        return selected_enterprise_name if selected_enterprise_name else ""
    else:
        return ""

def generate_field_info(field_name, remark, row_number, test_data, selected_enterprise_code, selected_enterprise_name):
    """
    生成字段信息，包含data和meta部分，并同步测试数据。
    """
    today = get_current_date()
    # Ensure test_value is synchronized for enterprise_code and enterprise_name fields
    test_value = get_test_value(field_name, test_data, selected_enterprise_code, selected_enterprise_name)
    
    field_info = {
        "field_name": field_name,
        "remark": remark,
        "row_number": row_number,
        "data": {
            "value": test_value,  # 设置为空字符串
            "relate_pic": "",     # 设置为空字符串
            "relate_doc": "",     # 设置为空字符串
            "relate_video": "",    # 设置为空字符串
            "data_url":f"http://get_{selected_enterprise_code}_data/{field_name}"
        },
        "meta": {
            "basic_meta": {
                "meta_content": f"{field_name}的基本描述性元数据",
                "identifier": "", # 设置为空字符串
                "title": f"{field_name}数据资源",
                "description": f"企业{field_name}相关数据的详细信息记录",
                "create_date": today,
                "update_date": today,
                "creator": "数据管理系统",
                "subject": f"{field_name},企业信息,基础数据"
            },
            "tech_meta": {
                "meta_content": f"{field_name}的技术性描述元数据",
                "format": "text/plain", # 默认为text/plain
                "size": "0 bytes",      # 设置为0 bytes
                "location": "",         # 设置为空字符串
                "version": "v1.0",
                "encoding": "UTF-8"
            },
            "manage_meta": {
                "meta_content": f"{field_name}的管理性字段描述元数据",
                "rights": "企业内部使用，受数据保护法规约束",
                "license": "企业数据使用许可",
                "source": "企业登记系统",
                "quality": "高质量数据，定期更新验证",
                "history": f"数据创建于 {today}，来源于官方企业登记信息"
            },
            "data_dic": {
                "meta_content": f"{field_name}字段对应的数据字典",
                "value1":"", # 设置为空字符串
                "value2":"", # 设置为空字符串
                "value3":""  # 设置为空字符串
            }
        }
    }
    return field_info

def read_excel_with_merged_cells(file_path, sheet_name=None, selected_enterprise_code=None, selected_enterprise_name=None):
    """
    读取包含合并单元格的Excel文件并转换为4级JSON结构
    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称，默认为None（第一个工作表）
        selected_enterprise_code: 选定的企业代码，用于同步
        selected_enterprise_name: 选定的企业名称，用于同步
    Returns:
        dict: 4级嵌套的JSON结构
    """
    # 使用openpyxl读取Excel文件以处理合并单元格
    workbook = load_workbook(file_path)
    if sheet_name is None:
        worksheet = workbook.active
    else:
        worksheet = workbook[sheet_name]
    
    # 获取所有数据
    data = []
    for row in worksheet.iter_rows(values_only=True):
        data.append(list(row))
    # 处理合并单元格 - 填充合并单元格的值
    merged_ranges = worksheet.merged_cells.ranges
    for merged_range in merged_ranges:
        # 获取合并区域的值（左上角单元格的值）
        merge_value = worksheet[merged_range.coord.split(':')[0]].value 
        # 填充合并区域内的所有单元格
        for row_idx in range(merged_range.min_row - 1, merged_range.max_row):
            for col_idx in range(merged_range.min_col - 1, merged_range.max_col):
                if row_idx < len(data) and col_idx < len(data[row_idx]):
                    data[row_idx][col_idx] = merge_value
    # 构建4级JSON结构
    result = {}
    # 假设表格从第2行开始（第1行是标题）
    for row_idx, row_data in enumerate(data[1:], start=2):  # 跳过标题行
        if len(row_data) < 5:  # 确保至少有5列数据
            continue     
        # 提取各级分类数据
        level1 = row_data[1] if row_data[1] is not None else ""  # 一级分类
        level2 = row_data[2] if row_data[2] is not None else ""  # 二级分类
        level3 = row_data[3] if row_data[3] is not None else ""  # 三级分类
        level4 = row_data[4] if row_data[4] is not None else ""  # 展示字段
        remark = row_data[5] if len(row_data) > 5 and row_data[5] is not None else ""  # 备注
        
        # 跳过空行
        if not any([level1, level2, level3, level4]):
            continue    
        # 构建层级结构
        if level1 not in result:
            result[level1] = {}
        
        if level2 not in result[level1]:
            result[level1][level2] = {}
        
        if level3 not in result[level1][level2]:
            result[level1][level2][level3] = []
        
        # 生成第四级数据（展示字段），传入选定的企业代码和名称进行同步
        field_info = generate_field_info(level4, remark, row_idx, test_data, selected_enterprise_code, selected_enterprise_name)
        result[level1][level2][level3].append(field_info)
    
    return result

def convert_to_enterprise_json(data, enterprise_code, enterprise_name):
    """
    转换为企业为主文档的JSON结构，并添加各级meta信息
    """
    children = []
    today = get_current_date()

    for level1_name, level1_data in data.items():
        level1_children = []
        for level2_name, level2_data in level1_data.items():
            level2_children = []
            for level3_name, level3_data in level2_data.items():
                level3_item = {
                    "name": level3_name,
                    "type": "level3",
                    "fields": level3_data,
                    "meta": {
                        "LevelContent": f"{level3_name}三级企业文档分类说明",
                        "Level4Nums": len(level3_data),
                        "CreateTime": today,
                        "UpdateTime": today,
                        "Description": f"包含{len(level3_data)}个具体数据字段的{level3_name}分类"
                    }
                }
                level2_children.append(level3_item)

            level2_item = {
                "name": level2_name,
                "type": "level2",
                "children": level2_children,
                "meta": {
                    "LevelContent": f"{level2_name}二级企业文档分类说明",
                    "Level3Nums": len(level2_children),
                    "CreateTime": today,
                    "UpdateTime": today,
                    "Description": f"包含{len(level2_children)}个三级企业文档的{level2_name}分类"
                }
            }
            level1_children.append(level2_item)

        level1_item = {
            "name": level1_name,
            "type": "level1",
            "children": level1_children,
            "meta": {
                "LevelContent": f"{level1_name}一级企业文档分类说明",
                "Level2Nums": len(level1_children),
                "CreateTime": today,
                "UpdateTime": today,
                "Description": f"包含{len(level1_children)}个二级企业文档的{level1_name}分类"
            }
        }
        children.append(level1_item)

    return {
        "EnterpriseCode": enterprise_code,
        "EnterpriseName": enterprise_name,
        "Level1Nums": len(children),
        "type": "level0",
        "CreateTime": today,
        "UpdateTime": today,
        "Description": f"企业：{enterprise_name} 档案信息，包含{len(children)}个企业子文档",
        "children": children
    }

def archive_json(enterprise_code, enterprise_name):
    """
    生成企业档案 JSON 数据
    
    Args:
        enterprise_code (str): 企业统一社会信用代码
        enterprise_name (str): 企业名称
    
    Returns:
        dict: 企业档案 JSON 数据，如果失败则返回错误信息字典
    """
    # 获取当前脚本所在目录的绝对路径
    import os
    current_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(current_dir, "一企一档数据项.xlsx")
    
    # 调试信息：打印文件路径和存在性
    print(f"当前工作目录: {os.getcwd()}")
    print(f"脚本所在目录: {current_dir}")
    print(f"Excel文件路径: {file_path}")
    print(f"文件是否存在: {os.path.exists(file_path)}")
    
    # 如果文件不存在，尝试在当前工作目录下查找
    if not os.path.exists(file_path):
        fallback_path = os.path.join(os.getcwd(), "一企一档数据项.xlsx")
        print(f"尝试备用路径: {fallback_path}")
        print(f"备用路径文件是否存在: {os.path.exists(fallback_path)}")
        if os.path.exists(fallback_path):
            file_path = fallback_path
        else:
            # 列出当前目录和脚本目录的文件，帮助调试
            print("当前工作目录下的文件:")
            try:
                for f in os.listdir(os.getcwd()):
                    if f.endswith('.xlsx'):
                        print(f"  - {f}")
            except:
                print("  无法列出当前目录文件")
            
            print("脚本所在目录下的文件:")
            try:
                for f in os.listdir(current_dir):
                    if f.endswith('.xlsx'):
                        print(f"  - {f}")
            except:
                print("  无法列出脚本目录文件")
    try:
        print("正在读取Excel文件...")
        # 不再从测试数据中随机选择企业，而是使用传入的参数
        enterprise_code = enterprise_code
        enterprise_name = enterprise_name
        # 将企业代码和名称传递给read_excel_with_merged_cells
        json_data = read_excel_with_merged_cells(file_path, selected_enterprise_code=enterprise_code, selected_enterprise_name=enterprise_name)
        # 转换为企业主文档格式
        enterprise_json = convert_to_enterprise_json(json_data, enterprise_code, enterprise_name)

        # 保存到文件（可选）
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        output_file = f"企业档案_结构化_{current_time}.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(enterprise_json, f, ensure_ascii=False, indent=2)
        print(f"转换完成！文件已保存为: {output_file}")
        print(f"企业名称: {enterprise_json['EnterpriseName']}")
        print(f"企业代码: {enterprise_json['EnterpriseCode']}")
        print(f"一级分类数量: {enterprise_json['Level1Nums']}")
        
        # 统计总字段数
        total_fields = 0
        for level1 in enterprise_json["children"]:
            for level2 in level1["children"]:
                for level3 in level2["children"]:
                    total_fields += len(level3["fields"])       
        print(f"总字段数量: {total_fields}")
        print("\n预览部分字段数据:")
        
        # 显示第一个字段的详细信息作为示例
        if enterprise_json["children"]:
            first_level1 = enterprise_json["children"][0]
            if first_level1["children"]:
                first_level2 = first_level1["children"][0]
                if first_level2["children"]:
                    first_level3 = first_level2["children"][0]
                    if first_level3["fields"]:
                        sample_field = first_level3["fields"][0]
                        print(f"示例字段: {sample_field['field_name']}")
                        print(f"测试值: {sample_field['data']['value']}")
                        print(f"字段描述: {sample_field['meta']['basic_meta']['description']}")
        
        # 确保返回字典类型
        return enterprise_json
        
    except FileNotFoundError:
        error_msg = f"错误: 找不到文件 '{file_path}'"
        print(error_msg)
        print(f"请确保文件 '一企一档数据项.xlsx' 存在于以下任一位置:")
        print(f"1. 脚本目录: {current_dir}")
        print(f"2. 当前工作目录: {os.getcwd()}")
        
        # 返回错误信息字典而不是 None
        return {
            "error": True,
            "error_type": "FileNotFoundError",
            "error_message": error_msg,
            "file_path_attempted": file_path,
            "script_directory": current_dir,
            "working_directory": os.getcwd(),
            "enterprise_code": enterprise_code,
            "enterprise_name": enterprise_name,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        error_msg = f"处理过程中出现错误: {str(e)}"
        print(error_msg)
        import traceback
        traceback.print_exc()
        # 返回错误信息字典而不是 None
        return {
            "error": True,
            "error_type": type(e).__name__,
            "error_message": error_msg,
            "enterprise_code": enterprise_code,
            "enterprise_name": enterprise_name,
            "timestamp": datetime.now().isoformat(),
            "traceback": traceback.format_exc()
        }


def archive_json_template(enterprise_code, enterprise_name):
    """
    生成企业档案 JSON 数据
    
    Args:
        enterprise_code (str): 企业统一社会信用代码
        enterprise_name (str): 企业名称
    
    Returns:
        dict: 企业档案 JSON 数据，如果失败则返回错误信息字典
    """
    # 获取当前脚本所在目录的绝对路径
    import os
    current_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(current_dir, "一企一档数据项.xlsx")
    
    # 调试信息：打印文件路径和存在性
    print(f"当前工作目录: {os.getcwd()}")
    print(f"脚本所在目录: {current_dir}")
    print(f"Excel文件路径: {file_path}")
    print(f"文件是否存在: {os.path.exists(file_path)}")
    
    # 如果文件不存在，尝试在当前工作目录下查找
    if not os.path.exists(file_path):
        fallback_path = os.path.join(os.getcwd(), "一企一档数据项.xlsx")
        print(f"尝试备用路径: {fallback_path}")
        print(f"备用路径文件是否存在: {os.path.exists(fallback_path)}")
        if os.path.exists(fallback_path):
            file_path = fallback_path
        else:
            # 列出当前目录和脚本目录的文件，帮助调试
            print("当前工作目录下的文件:")
            try:
                for f in os.listdir(os.getcwd()):
                    if f.endswith('.xlsx'):
                        print(f"  - {f}")
            except:
                print("  无法列出当前目录文件")
            
            print("脚本所在目录下的文件:")
            try:
                for f in os.listdir(current_dir):
                    if f.endswith('.xlsx'):
                        print(f"  - {f}")
            except:
                print("  无法列出脚本目录文件")
    try:
        print("正在读取Excel文件...")
        # 不再从测试数据中随机选择企业，而是使用传入的参数
        enterprise_code = enterprise_code
        enterprise_name = enterprise_name

        # 将企业代码和名称传递给read_excel_with_merged_cells
        json_data = read_excel_with_merged_cells(file_path, selected_enterprise_code=enterprise_code, selected_enterprise_name=enterprise_name)
        # 转换为企业主文档格式
        enterprise_json = convert_to_enterprise_json(json_data, enterprise_code, enterprise_name)
        # 保存到文件（可选）
        # current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        # output_file = f"企业档案_结构化_{current_time}.json"
        # with open(output_file, 'w', encoding='utf-8') as f:
        #     json.dump(enterprise_json, f, ensure_ascii=False, indent=2)

        # print(f"转换完成！文件已保存为: {output_file}")
        print(f"企业名称: {enterprise_json['EnterpriseName']}")
        print(f"企业代码: {enterprise_json['EnterpriseCode']}")
        print(f"一级分类数量: {enterprise_json['Level1Nums']}")
        
        # 统计总字段数
        total_fields = 0
        for level1 in enterprise_json["children"]:
            for level2 in level1["children"]:
                for level3 in level2["children"]:
                    total_fields += len(level3["fields"])       
        print(f"总字段数量: {total_fields}")
        print("\n预览部分字段数据:")
        
        # 显示第一个字段的详细信息作为示例
        if enterprise_json["children"]:
            first_level1 = enterprise_json["children"][0]
            if first_level1["children"]:
                first_level2 = first_level1["children"][0]
                if first_level2["children"]:
                    first_level3 = first_level2["children"][0]
                    if first_level3["fields"]:
                        sample_field = first_level3["fields"][0]
                        print(f"示例字段: {sample_field['field_name']}")
                        print(f"测试值: {sample_field['data']['value']}")
                        print(f"字段描述: {sample_field['meta']['basic_meta']['description']}")
        
        # 确保返回字典类型
        return enterprise_json
        
    except FileNotFoundError:
        error_msg = f"错误: 找不到文件 '{file_path}'"
        print(error_msg)
        print(f"请确保文件 '一企一档数据项.xlsx' 存在于以下任一位置:")
        print(f"1. 脚本目录: {current_dir}")
        print(f"2. 当前工作目录: {os.getcwd()}")
        
        # 返回错误信息字典而不是 None
        return {
            "error": True,
            "error_type": "FileNotFoundError",
            "error_message": error_msg,
            "file_path_attempted": file_path,
            "script_directory": current_dir,
            "working_directory": os.getcwd(),
            "enterprise_code": enterprise_code,
            "enterprise_name": enterprise_name,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        error_msg = f"处理过程中出现错误: {str(e)}"
        print(error_msg)
        import traceback
        traceback.print_exc()
        # 返回错误信息字典而不是 None
        return {
            "error": True,
            "error_type": type(e).__name__,
            "error_message": error_msg,
            "enterprise_code": enterprise_code,
            "enterprise_name": enterprise_name,
            "timestamp": datetime.now().isoformat(),
            "traceback": traceback.format_exc()
        }
    

# if __name__ == "__main__":
#     archive_json_template("91110000MA01ATCD15","青岛科技发展有限公司")