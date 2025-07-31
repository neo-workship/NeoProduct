from openpyxl import load_workbook
from datetime import datetime
from typing import List, Dict, Any, Optional
import hashlib

class FlatEnterpriseArchiveGenerator:
    """
    扁平化企业档案生成器
    将Excel的层级结构转换为扁平化的MongoDB文档集合
    """
    
    def __init__(self):
        self.level_counters = {
            'l1': {},  # 一级分类计数器
            'l2': {},  # 二级分类计数器  
            'l3': {},  # 三级分类计数器
            'field': {}  # 字段计数器
        }
        
    def generate_code(self, name: str, level: str, parent_code: str = "") -> str:
        """
        生成层级编码
        Args:
            name: 分类或字段名称
            level: 层级类型 (l1, l2, l3, field)
            parent_code: 父级编码
        Returns:
            str: 生成的编码
        """
        # 使用MD5哈希确保编码唯一性和一致性
        hash_input = f"{parent_code}_{name}" if parent_code else name
        hash_code = hashlib.md5(hash_input.encode('utf-8')).hexdigest()[:6].upper()
        
        if level == 'l1':
            return f"L1{hash_code}"
        elif level == 'l2':
            return f"L2{hash_code}"
        elif level == 'l3':
            return f"L3{hash_code}"
        elif level == 'field':
            return f"F{hash_code}"
        
        return hash_code

    def read_excel_with_merged_cells(self, file_path: str, 
                                     sheet_name: Optional[str] = None,
                                     enterprise_code: str = "", 
                                     enterprise_name: str = "") -> List[Dict[str, Any]]:
        """
        读取包含合并单元格的Excel文件并转换为扁平化文档列表
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称，默认为None（第一个工作表）
        Returns:
            List[Dict]: 扁平化的字段文档列表
        """
        # 使用openpyxl读取Excel文件以处理合并单元格
        workbook = load_workbook(file_path)
        if sheet_name is None:
            worksheet = workbook.active
        else:
            worksheet = workbook[sheet_name]
            
        # 获取所有数据
        data = []
        for row in worksheet.iter_rows(values_only=True):
            data.append(list(row))
            
        # 处理合并单元格 - 填充合并单元格的值
        merged_ranges = worksheet.merged_cells.ranges
        for merged_range in merged_ranges:
            # 获取合并区域的值（左上角单元格的值）
            merge_value = worksheet[merged_range.coord.split(':')[0]].value 
            # 填充合并区域内的所有单元格
            for row_idx in range(merged_range.min_row - 1, merged_range.max_row):
                for col_idx in range(merged_range.min_col - 1, merged_range.max_col):
                    if row_idx < len(data) and col_idx < len(data[row_idx]):
                        data[row_idx][col_idx] = merge_value
                        
        # 转换为扁平化文档列表
        flat_documents = []
        current_time = datetime.now()
        
        # 用于跟踪当前的层级信息
        current_l1 = {"name": "", "code": ""}
        current_l2 = {"name": "", "code": ""}
        current_l3 = {"name": "", "code": ""}
        
        # 用于排序
        l1_order = 0
        l2_order = 0
        l3_order = 0
        
        # 假设表格从第2行开始（第1行是标题）
        for row_idx, row_data in enumerate(data[1:], start=2):
            if len(row_data) < 5:  # 确保至少有5列数据
                continue
                
            # 提取各级分类数据
            level1 = row_data[1] if row_data[1] is not None else ""
            level2 = row_data[2] if row_data[2] is not None else ""
            level3 = row_data[3] if row_data[3] is not None else ""
            field_name = row_data[4] if row_data[4] is not None else ""
            remark = row_data[5] if len(row_data) > 5 and row_data[5] is not None else ""
            data_url = row_data[6] if len(row_data) > 6 and row_data[6] is not None else ""
            
            # 跳过空行
            if not any([level1, level2, level3, field_name]):
                continue
                
            # 更新层级信息和编码
            if level1 and level1 != current_l1["name"]:
                l1_order += 1
                l2_order = 0
                l3_order = 0
                current_l1 = {
                    "name": level1,
                    "code": self.generate_code(level1, "l1")
                }
                
            if level2 and level2 != current_l2["name"]:
                l2_order += 1
                l3_order = 0
                current_l2 = {
                    "name": level2,
                    "code": self.generate_code(level2, "l2", current_l1["code"])
                }
                
            if level3 and level3 != current_l3["name"]:
                l3_order += 1
                current_l3 = {
                    "name": level3,
                    "code": self.generate_code(level3, "l3", current_l2["code"])
                }
            
            # 生成字段编码
            field_code = self.generate_code(field_name, "field", current_l3["code"])
            
            # 构建路径
            path_code = f"{current_l1['code']}.{current_l2['code']}.{current_l3['code']}"
            path_name = f"{current_l1['name']}.{current_l2['name']}.{current_l3['name']}"
            full_path_code = f"{path_code}.{field_code}"
            full_path_name = f"{path_name}.{field_name}"
            
            # 生成扁平化文档
            field_document = {
                # 企业信息
                "enterprise_code": f"{enterprise_code}",
                "enterprise_name": f"{enterprise_name}",
                
                # 层级信息
                "l1_code": current_l1["code"],
                "l1_name": current_l1["name"],
                "l2_code": current_l2["code"],
                "l2_name": current_l2["name"],
                "l3_code": current_l3["code"],
                "l3_name": current_l3["name"],
                
                # 路径信息
                "path_code": path_code,
                "path_name": path_name,
                "full_path_code": full_path_code,
                "full_path_name": full_path_name,
                
                # 字段信息
                "field_code": field_code,
                "field_name": field_name,
                "field_type": self._infer_field_type(field_name),
                
                # 数据值
                "value": "",          # 初始值为空
                "value_text": "",     # 用于全文检索的文本值
                "value_pic_url":"",   # 值对应图片url，指向文件服务器
                "value_doc_url":"",   # 值对应文档url，指向文件服务器
                "value_video_url":"", # 值对应视频url，指向文件服务器

                # 元数据
                "remark": remark,
                "data_url": data_url,
                "is_required": self._is_required_field(field_name, remark),
                "data_source": self._extract_data_source(remark),
                "encoding": "UTF-8",
                "format": "text/plain",
                "license": "企业数据使用许可",
                "rights": "企业内部使用，受数据保护法规约束",
                "update_frequency":"更新频率",
                "value_dict":"{xx,xx,xx}",
                
                # 排序和显示
                "l1_order": l1_order,
                "l2_order": l2_order,
                "l3_order": l3_order,
                "field_order": len(flat_documents) + 1,
                
                # 时间戳
                "create_time": current_time,
                "update_time": current_time,
                
                # 数据状态
                "status": "active"
            }
            
            flat_documents.append(field_document)
            
        return flat_documents
    
    def _infer_field_type(self, field_name: str) -> str:
        """
        根据字段名推断字段类型
        """
        field_name_lower = field_name.lower()
        
        if any(keyword in field_name_lower for keyword in ['金额', '资金', '资本', "总额", '收入', '支出', '数量', '面积']):
            return "number"
        elif any(keyword in field_name_lower for keyword in ['时间', '日期', '年份', '月份']):
            return "date"
        elif any(keyword in field_name_lower for keyword in ['是否', '有无']):
            return "boolean"
        elif any(keyword in field_name_lower for keyword in ['邮箱', '邮件']):
            return "email"
        elif any(keyword in field_name_lower for keyword in ['电话', '手机', '传真']):
            return "phone"
        elif any(keyword in field_name_lower for keyword in ['网址', '网站', 'url']):
            return "url"
        else:
            return "string"
    
    def _is_required_field(self, field_name: str, remark: str) -> bool:
        """
        判断字段是否为必填项
        """
        required_keywords = ['必填', '必须', '必需', '必要', '必备']
        field_text = f"{field_name} {remark}".lower()
        return any(keyword in field_text for keyword in required_keywords)
    
    def _extract_data_source(self, remark: str) -> str:
        """
        从备注中提取数据来源
        """
        if not remark:
            return ""
            
        sources = ['工商局', '税务局', '银行', '统计局', '人社局', '公安局', '海关']
        for source in sources:
            if source in remark:
                return source
        return ""

    def generate_hierarchy_structure_json(self, file_path: str, 
                                      output_path: str = "./hierarchy_structure.json",
                                      sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        基于read_excel_with_merged_cells改造的函数，专门用于生成层级结构JSON数据
        
        Args:
            file_path: Excel文件路径
            output_path: 输出JSON文件路径
            sheet_name: 工作表名称，默认为None（第一个工作表）
        
        Returns:
            Dict[str, Any]: 层级结构数据
        """
        from pathlib import Path
        import json
        
        # 使用openpyxl读取Excel文件以处理合并单元格
        workbook = load_workbook(file_path)
        if sheet_name is None:
            worksheet = workbook.active
        else:
            worksheet = workbook[sheet_name]
            
        # 获取所有数据
        data = []
        for row in worksheet.iter_rows(values_only=True):
            data.append(list(row))
            
        # 处理合并单元格 - 填充合并单元格的值
        merged_ranges = worksheet.merged_cells.ranges
        for merged_range in merged_ranges:
            merge_value = worksheet[merged_range.coord.split(':')[0]].value 
            for row_idx in range(merged_range.min_row - 1, merged_range.max_row):
                for col_idx in range(merged_range.min_col - 1, merged_range.max_col):
                    if row_idx < len(data) and col_idx < len(data[row_idx]):
                        data[row_idx][col_idx] = merge_value
        
        # 初始化层级结构数据
        hierarchy_structure = {
            "l1_categories": [],
            "metadata": {
                "total_l1": 0,
                "total_l2": 0,
                "total_l3": 0,
                "total_fields": 0,
                "generated_time": datetime.now().isoformat(),
                "version": "1.0",
                "source_file": file_path
            }
        }
        
        # 用于构建层级结构的字典和跟踪变量
        l1_dict = {}  # l1_code -> l1_data
        l2_dict = {}  # l2_code -> l2_data  
        l3_dict = {}  # l3_code -> l3_data
        
        current_l1 = {"name": "", "code": ""}
        current_l2 = {"name": "", "code": ""}
        current_l3 = {"name": "", "code": ""}
        
        l1_order = 0
        l2_order = 0
        l3_order = 0
        field_count = 0
        
        # 从第2行开始处理数据（第1行是标题）
        for row_idx, row_data in enumerate(data[1:], start=2):
            if len(row_data) < 5:
                continue
            
            # 提取各级分类数据
            level1 = row_data[1] if row_data[1] is not None else ""
            level2 = row_data[2] if row_data[2] is not None else ""
            level3 = row_data[3] if row_data[3] is not None else ""
            field_name = row_data[4] if row_data[4] is not None else ""
            data_url = row_data[6] if len(row_data) > 6 and row_data[6] is not None else ""

            # 跳过空行
            if not any([level1, level2, level3, field_name]):
                continue
            
            # 处理一级分类
            if level1 and level1 != current_l1["name"]:
                current_l1 = {
                    "name": level1,
                    "code": self.generate_code(level1, "l1")
                }
                
                if current_l1["code"] not in l1_dict:
                    l1_data = {
                        "l1_code": current_l1["code"],
                        "l1_name": current_l1["name"],
                        "l2_categories": [],
                        "path_code": f"{current_l1['code']}",
                        "path_name": f"{current_l1['name']}"
                    }
                    l1_dict[current_l1["code"]] = l1_data
                    hierarchy_structure["l1_categories"].append(l1_data)
            
            # 处理二级分类
            if level2 and level2 != current_l2["name"]:
                l2_order += 1
                l3_order = 0
                current_l2 = {
                    "name": level2,
                    "code": self.generate_code(level2, "l2", current_l1["code"])
                }
                
                if current_l2["code"] not in l2_dict:
                    l2_data = {
                        "l2_code": current_l2["code"],
                        "l2_name": current_l2["name"],
                        "l3_categories": [],
                        "path_code": f"{current_l1['code']}.{current_l2['code']}",
                        "path_name": f"{current_l1['name']}.{current_l2['name']}"
                    }
                    l2_dict[current_l2["code"]] = l2_data
                    l1_dict[current_l1["code"]]["l2_categories"].append(l2_data)
            
            # 处理三级分类
            if level3 and level3 != current_l3["name"]:
                l3_order += 1
                current_l3 = {
                    "name": level3,
                    "code": self.generate_code(level3, "l3", current_l2["code"])
                }
                
                if current_l3["code"] not in l3_dict:
                    path_code = f"{current_l1['code']}.{current_l2['code']}.{current_l3['code']}"
                    path_name = f"{current_l1['name']}.{current_l2['name']}.{current_l3['name']}"
                    
                    l3_data = {
                        "l3_code": current_l3["code"],
                        "l3_name": current_l3["name"],
                        "fields": [],
                        "path_code": path_code,
                        "path_name": path_name
                    }
                    l3_dict[current_l3["code"]] = l3_data
                    l2_dict[current_l2["code"]]["l3_categories"].append(l3_data)
            
            # 处理字段
            if field_name:
                field_count += 1
                field_code = self.generate_code(field_name, "field", current_l3["code"])
                
                path_code = f"{current_l1['code']}.{current_l2['code']}.{current_l3['code']}"
                path_name = f"{current_l1['name']}.{current_l2['name']}.{current_l3['name']}"
                full_path_code = f"{path_code}.{field_code}"
                full_path_name = f"{path_name}.{field_name}"
                
                field_data = {
                    "field_code": field_code,
                    "field_name": field_name,
                    "full_path_code": full_path_code,
                    "full_path_name": full_path_name
                }
                l3_dict[current_l3["code"]]["fields"].append(field_data)
        
        # 更新统计信息
        hierarchy_structure["metadata"]["total_l1"] = len(l1_dict)
        hierarchy_structure["metadata"]["total_l2"] = len(l2_dict)
        hierarchy_structure["metadata"]["total_l3"] = len(l3_dict)
        hierarchy_structure["metadata"]["total_fields"] = field_count
        
        # 保存JSON文件
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(hierarchy_structure, f, ensure_ascii=False, indent=2)
        
        print(f"✅ 层级结构JSON文件已生成: {output_path}")
        return hierarchy_structure

def generate_doc(enterprise_code: str = "", enterprise_name: str = ""):
    """
    主函数：演示扁平化文档生成流程
    """
    # 创建生成器实例
    generator = FlatEnterpriseArchiveGenerator()
    
    # Excel文件路径
    file_path = "./一企一档数据项.xlsx"
    
    try:
        print("开始读取Excel文件...")
        # 读取Excel并生成模板文档
        template_documents = generator.read_excel_with_merged_cells(
            file_path,
            enterprise_code=enterprise_code, 
            enterprise_name=enterprise_name
        )
        print(f"成功生成 {len(template_documents)} 个字段模板")
        return template_documents
        
    except FileNotFoundError:
        print(f"错误: 找不到文件 {file_path}")
        print("请确保Excel文件存在且路径正确")
        return {}
    except Exception as e:
        print(f"错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return {}

def create_hierarchy_json(excel_file: str = "./一企一档数据项.xlsx", 
                         json_file: str = "./hierarchy_structure.json") -> Dict[str, Any]:
    """
    便捷函数：生成层级结构JSON文件
    
    Args:
        excel_file: Excel文件路径
        json_file: 输出JSON文件路径
    
    Returns:
        Dict[str, Any]: 层级结构数据
    """
    generator = FlatEnterpriseArchiveGenerator()
    return generator.generate_hierarchy_structure_json(excel_file, json_file)

if __name__ == "__main__":
    create_hierarchy_json()